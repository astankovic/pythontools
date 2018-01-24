from os import path
from os import listdir
from openpyxl import Workbook
from openpyxl import load_workbook
import re

path._getfullpathname

def read_the_folder_for_xlsx():
    xlsx_files = list()
    file_list = listdir(path.curdir)
    for i in file_list:
        if path.isfile(i) & (path.splitext(i)[1] == ".xlsx"):
            xlsx_files.append(path.abspath(i))
    return xlsx_files

def vlookup_across_multiple_xls(term, search_column, rel_lookup_column, include_file_name_in_output):
    result = list()
    for r in read_the_folder_for_xlsx():
        wb = load_workbook(r)
        for sheet in wb:
            for i in range(1,sheet.max_row + 1):
                iterator = sheet.cell(row = i, column = search_column)
                if iterator.value == term:
                    result_piece = [sheet.cell(row = i, column = search_column + rel_lookup_column).value, sheet.title, path.basename(r)]
                    result.append(result_piece)
                    print('{}\t{}\t{}'.format(sheet.cell(row = i, column = search_column + rel_lookup_column).value, sheet.title, path.basename(r)))
    return result

def export_results_to_csv(data, dest_filename):
    with open(dest_filename + '.csv', 'a') as the_file:
        for result in data:
            separat = ','
            the_file.write(separat.join(str(x) for x in result) + '\n')
    
def export_results_to_xlsx(data, dest_filename):
    resultswb = Workbook()
    sheet1 = resultswb.active
    sheet1.title = "vlookup results"
    

    resultswb.save(filename = dest_filename)

searchText = input("Search term: ")
result = vlookup_across_multiple_xls(searchText, 2, 2, False)
export_results_to_csv(result, path._getfullpathname(path.curdir) + '\\vlookup_result')
export_results_to_xlsx(result, path._getfullpathname(path.curdir) + '\\vlookup_result')
text = input("Good with this?")